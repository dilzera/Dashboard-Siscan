import sys
import os
import unicodedata
import re

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.data_layer import get_engine
from src.models import User, Base, get_session
from sqlalchemy import text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def normalize_username(name):
    nfkd = unicodedata.normalize('NFKD', name)
    ascii_name = nfkd.encode('ASCII', 'ignore').decode('ASCII')
    username = ascii_name.lower().strip()
    username = re.sub(r'[^a-z0-9]+', '.', username)
    username = username.strip('.')
    return username


def normalize_password(name):
    nfkd = unicodedata.normalize('NFKD', name)
    ascii_name = nfkd.encode('ASCII', 'ignore').decode('ASCII')
    pwd_base = ascii_name.lower().replace(' ', '')
    return f"{pwd_base}@2026"


def get_district_for_unit(engine, unit_name):
    with engine.connect() as conn:
        result = conn.execute(
            text("SELECT DISTINCT distrito_sanitario FROM exam_records WHERE unidade_de_saude__nome = :unit LIMIT 1"),
            {"unit": unit_name}
        )
        row = result.fetchone()
        return row[0] if row else None


def get_district_for_provider(engine, provider_name):
    with engine.connect() as conn:
        result = conn.execute(
            text("SELECT DISTINCT distrito_sanitario FROM exam_records WHERE prestador_de_servico__nome = :prov LIMIT 1"),
            {"prov": provider_name}
        )
        row = result.fetchone()
        return row[0] if row else None


def main():
    engine = get_engine()
    db_session = get_session()

    with engine.connect() as conn:
        units = [r[0] for r in conn.execute(
            text("SELECT DISTINCT unidade_de_saude__nome FROM exam_records WHERE unidade_de_saude__nome IS NOT NULL ORDER BY unidade_de_saude__nome")
        ).fetchall()]

        districts = [r[0] for r in conn.execute(
            text("SELECT DISTINCT distrito_sanitario FROM exam_records WHERE distrito_sanitario IS NOT NULL ORDER BY distrito_sanitario")
        ).fetchall()]

        providers = [r[0] for r in conn.execute(
            text("SELECT DISTINCT prestador_de_servico__nome FROM exam_records WHERE prestador_de_servico__nome IS NOT NULL ORDER BY prestador_de_servico__nome")
        ).fetchall()]

    users_data = []

    users_data.append({
        'tipo': 'Secretaria',
        'referencia': 'Secretaria Municipal de Saúde',
        'username': 'secretaria.saude',
        'password': 'secretariasaude@2026',
        'role': 'admin',
        'access_level': 'secretaria',
        'district': None,
        'health_unit': None,
        'name': 'Secretaria Municipal de Saúde'
    })

    for district in districts:
        username = f"distrito.{normalize_username(district)}"
        password = f"distrito{normalize_username(district).replace('.', '')}@2026"
        users_data.append({
            'tipo': 'Distrito',
            'referencia': district,
            'username': username,
            'password': password,
            'role': 'admin',
            'access_level': 'distrito',
            'district': district,
            'health_unit': None,
            'name': f'Gestor {district}'
        })

    for unit in units:
        username = normalize_username(unit)
        password = normalize_password(unit)
        district = get_district_for_unit(engine, unit)
        users_data.append({
            'tipo': 'Unidade de Saúde',
            'referencia': unit,
            'username': username,
            'password': password,
            'role': 'viewer',
            'access_level': 'unidade',
            'district': district,
            'health_unit': unit,
            'name': unit
        })

    for provider in providers:
        username = f"prestador.{normalize_username(provider)}"
        password = f"prestador{normalize_username(provider).replace('.', '')}@2026"
        district = get_district_for_provider(engine, provider)
        users_data.append({
            'tipo': 'Prestador',
            'referencia': provider,
            'username': username,
            'password': password,
            'role': 'viewer',
            'access_level': 'unidade',
            'district': district,
            'health_unit': provider,
            'name': provider
        })

    created = 0
    skipped = 0
    for ud in users_data:
        existing = db_session.query(User).filter_by(username=ud['username']).first()
        if existing:
            print(f"  SKIP: {ud['username']} (ja existe)")
            skipped += 1
            continue

        user = User(
            username=ud['username'],
            name=ud['name'],
            role=ud['role'],
            access_level=ud['access_level'],
            district=ud['district'],
            health_unit=ud['health_unit'],
            is_active=True,
            must_change_password=False
        )
        user.set_password(ud['password'])
        db_session.add(user)
        created += 1
        print(f"  OK: {ud['username']} ({ud['tipo']})")

    db_session.commit()
    db_session.close()
    print(f"\nTotal: {created} criados, {skipped} ignorados (ja existiam)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios POC"

    headers = ['Tipo', 'Referência', 'Usuário', 'Senha', 'Nível de Acesso', 'Distrito', 'Unidade de Saúde']
    header_fill = PatternFill(start_color='17A2B8', end_color='17A2B8', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    tipo_colors = {
        'Secretaria': PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid'),
        'Distrito': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
        'Unidade de Saúde': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
        'Prestador': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
    }

    for row_idx, ud in enumerate(users_data, 2):
        row_data = [
            ud['tipo'],
            ud['referencia'],
            ud['username'],
            ud['password'],
            ud['access_level'],
            ud['district'] or '',
            ud['health_unit'] or ''
        ]
        fill = tipo_colors.get(ud['tipo'])
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if fill:
                cell.fill = fill

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 55

    ws.auto_filter.ref = f"A1:G{len(users_data) + 1}"

    output_path = 'usuarios_poc_siscan.xlsx'
    wb.save(output_path)
    print(f"\nPlanilha gerada: {output_path}")
    print(f"Total de linhas: {len(users_data)}")


if __name__ == '__main__':
    main()
